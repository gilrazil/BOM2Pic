from __future__ import annotations

import io
import posixpath
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PIL import Image


NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}

REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass
class AnchoredImage:
    row: int  # zero-based
    col: int  # zero-based
    media_path: str  # e.g., xl/media/image1.png


@dataclass
class ExtractedImageDetail:
    """Detailed extracted image record.

    row: 1-based Excel row index
    sheet: worksheet title
    name_raw: cell value from the name column (unmodified string)
    image_bytes: original image bytes from the workbook (unconverted)
    """
    row: int
    sheet: str
    name_raw: str
    image_bytes: bytes


def column_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (A..Z, AA, AB..) to zero-based index."""
    return int(column_index_from_string(letter.strip().upper())) - 1


def _resolve_rel_target(rels_file_path: str, target: str) -> str:
    """Resolve a relationship Target relative to its SOURCE part, not the _rels folder.

    For a rels path like xl/worksheets/_rels/sheet1.xml.rels, the source part is
    xl/worksheets/sheet1.xml. Relationship targets are resolved relative to the
    source part's directory (xl/worksheets/), so we must move one directory up
    from the _rels directory before joining the target.
    """
    source_base_dir = posixpath.dirname(posixpath.dirname(rels_file_path))
    return posixpath.normpath(posixpath.join(source_base_dir, target))


def _get_first_sheet_paths(zf: zipfile.ZipFile) -> Tuple[str, str]:
    """Return (sheet_xml_path, sheet_rels_path) for the first worksheet.

    Determines the first sheet from `xl/workbook.xml` and resolves via
    `xl/_rels/workbook.xml.rels` instead of assuming `sheet1.xml`.
    """
    workbook_xml = "xl/workbook.xml"
    workbook_rels = "xl/_rels/workbook.xml.rels"
    if workbook_xml not in zf.namelist():
        raise ValueError("Invalid workbook: missing xl/workbook.xml")
    if workbook_rels not in zf.namelist():
        raise ValueError("Invalid workbook: missing xl/_rels/workbook.xml.rels")

    wb_root = ET.fromstring(zf.read(workbook_xml))
    first_sheet_el = wb_root.find(".//s:sheets/s:sheet", NS)
    if first_sheet_el is None:
        raise ValueError("No worksheets found in workbook")
    r_id = first_sheet_el.attrib.get(f"{{{NS['r']}}}id")
    if not r_id:
        raise ValueError("First sheet missing relationship id")

    rels_root = ET.fromstring(zf.read(workbook_rels))
    sheet_target: str | None = None
    # Prefer exact match by Id
    for rel in rels_root.findall(f".//{{{REL_NS}}}Relationship"):
        if rel.attrib.get("Id") == r_id:
            sheet_target = rel.attrib.get("Target")
            break
    # Fallback: first worksheet relationship
    if not sheet_target:
        for rel in rels_root.findall(f".//{{{REL_NS}}}Relationship"):
            rel_type = rel.attrib.get("Type", "")
            if rel_type.endswith("/worksheet"):
                sheet_target = rel.attrib.get("Target")
                break
    if not sheet_target:
        raise ValueError("Could not resolve first sheet target")

    sheet_xml = _resolve_rel_target(workbook_rels, sheet_target)
    sheet_rels = posixpath.join(posixpath.dirname(sheet_xml), "_rels", posixpath.basename(sheet_xml) + ".rels")
    return sheet_xml, sheet_rels


def _find_drawing_for_sheet(zf: zipfile.ZipFile, sheet_xml: str, sheet_rels: str) -> str | None:
    """Return drawing XML path for the sheet, or None if no drawings."""
    if sheet_rels not in zf.namelist():
        return None

    # Find drawing r:id in sheet xml
    sheet_root = ET.fromstring(zf.read(sheet_xml))
    drawing_el = sheet_root.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing")
    if drawing_el is None:
        return None
    r_id = drawing_el.attrib.get(f"{{{NS['r']}}}id")
    if not r_id:
        return None

    # Map r:id to target drawing path via relationships
    rels_root = ET.fromstring(zf.read(sheet_rels))
    for rel in rels_root.findall(f".//{{{REL_NS}}}Relationship"):
        if rel.attrib.get("Id") == r_id:
            target = rel.attrib.get("Target")
            if not target:
                continue
            return _resolve_rel_target(sheet_rels, target)
    return None


def _map_drawing_relations(zf: zipfile.ZipFile, drawing_xml_path: str) -> Dict[str, str]:
    """Return map of r:embed id -> media path for a drawing file."""
    rels_path = posixpath.join(posixpath.dirname(drawing_xml_path), "_rels", posixpath.basename(drawing_xml_path) + ".rels")
    if rels_path not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_path))
    rmap: Dict[str, str] = {}
    for rel in root.findall(f".//{{{REL_NS}}}Relationship"):
        r_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if r_id and target:
            # Targets here are relative to the drawing part directory
            media_path = _resolve_rel_target(rels_path, target)
            rmap[r_id] = media_path
    return rmap


def _parse_anchored_images(zf: zipfile.ZipFile, drawing_xml_path: str) -> List[Tuple[int, int, str]]:
    """Parse drawing xml and return list of (row, col, r:embed)."""
    root = ET.fromstring(zf.read(drawing_xml_path))
    anchors = []

    # twoCellAnchor
    for node in root.findall(".//xdr:twoCellAnchor", NS):
        from_node = node.find("xdr:from", NS)
        pic_node = node.find("xdr:pic", NS)
        if from_node is None or pic_node is None:
            continue
        col_text = from_node.findtext("xdr:col", default="0", namespaces=NS)
        row_text = from_node.findtext("xdr:row", default="0", namespaces=NS)
        blip = pic_node.find("xdr:blipFill/a:blip", NS)
        if blip is None:
            continue
        r_embed = blip.attrib.get(f"{{{NS['r']}}}embed")
        if r_embed is None:
            continue
        anchors.append((int(row_text), int(col_text), r_embed))

    # oneCellAnchor
    for node in root.findall(".//xdr:oneCellAnchor", NS):
        from_node = node.find("xdr:from", NS)
        pic_node = node.find("xdr:pic", NS)
        if from_node is None or pic_node is None:
            continue
        col_text = from_node.findtext("xdr:col", default="0", namespaces=NS)
        row_text = from_node.findtext("xdr:row", default="0", namespaces=NS)
        blip = pic_node.find("xdr:blipFill/a:blip", NS)
        if blip is None:
            continue
        r_embed = blip.attrib.get(f"{{{NS['r']}}}embed")
        if r_embed is None:
            continue
        anchors.append((int(row_text), int(col_text), r_embed))

    return anchors


def _convert_to_png(image_bytes: bytes) -> bytes:
    with Image.open(io.BytesIO(image_bytes)) as img:
        img = img.convert("RGBA") if img.mode in ("LA", "P", "RGBA") else img.convert("RGB")
        out = io.BytesIO()
        img.save(out, format="PNG")
        return out.getvalue()


def _sanitize_filename(name: str) -> str:
    name = name.strip()
    # Replace invalid filename characters with underscore
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
    # Avoid empty names
    return name or "image"


def extract_images_by_column(
    xlsx_bytes: bytes,
    image_col_letter: str,
    name_col_letter: str,
    max_images: int | None = None,
) -> List[Tuple[str, bytes]]:
    """Extract images anchored in the specified image column and rename from name column.

    If ``max_images`` is ``None`` or less than or equal to zero, no per-file image limit is enforced.

    Returns list of tuples (filename, png_bytes) ordered by row.
    """
    img_col_idx = column_letter_to_index(image_col_letter)
    name_col_idx = column_letter_to_index(name_col_letter)

    # Load workbook for reading names
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    ws = wb.worksheets[0]

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        sheet_xml, sheet_rels = _get_first_sheet_paths(zf)
        drawing_xml_path = _find_drawing_for_sheet(zf, sheet_xml, sheet_rels)
        if not drawing_xml_path:
            return []

        rel_map = _map_drawing_relations(zf, drawing_xml_path)
        anchors = _parse_anchored_images(zf, drawing_xml_path)

        anchored_images: List[AnchoredImage] = []
        for row_zero, col_zero, r_embed in anchors:
            if r_embed in rel_map:
                media_path = rel_map[r_embed]
                anchored_images.append(AnchoredImage(row=row_zero, col=col_zero, media_path=media_path))

        # Filter by image column
        filtered = [ai for ai in anchored_images if ai.col == img_col_idx]

        # Sort by row for deterministic order
        filtered.sort(key=lambda x: x.row)

        results: List[Tuple[str, bytes]] = []
        used_names: Dict[str, int] = {}
        for ai in filtered:
            # Excel rows are 1-based
            row_number = ai.row + 1
            name_cell = ws.cell(row=row_number, column=name_col_idx + 1).value
            base_name = _sanitize_filename(str(name_cell) if name_cell is not None else f"image_row{row_number}")

            # Ensure unique filename
            counter = used_names.get(base_name, 0)
            used_names[base_name] = counter + 1
            filename = f"{base_name}.png" if counter == 0 else f"{base_name}_{counter}.png"

            # Read image bytes and convert to PNG
            raw_bytes = zf.read(ai.media_path)
            png_bytes = _convert_to_png(raw_bytes)
            results.append((filename, png_bytes))

            if max_images and max_images > 0 and len(results) > max_images:
                raise ValueError(f"Too many images: limit is {max_images}")

        return results


def extract_images_details(
    xlsx_bytes: bytes,
    image_col_letter: str,
    name_col_letter: str,
    max_images: int | None = None,
) -> List[ExtractedImageDetail]:
    """Extract raw images and associated metadata from the first worksheet.

    - Returns a list ordered top-to-bottom by row.
    - Does not modify image format; returns original bytes.
    - Reads the name column cell value as-is (converted to string later by caller).
    - If ``max_images`` is set (>0), enforces a cap per workbook.
    """
    img_col_idx = column_letter_to_index(image_col_letter)
    name_col_idx = column_letter_to_index(name_col_letter)

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    ws = wb.worksheets[0]

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        sheet_xml, sheet_rels = _get_first_sheet_paths(zf)
        drawing_xml_path = _find_drawing_for_sheet(zf, sheet_xml, sheet_rels)
        if not drawing_xml_path:
            return []

        rel_map = _map_drawing_relations(zf, drawing_xml_path)
        anchors = _parse_anchored_images(zf, drawing_xml_path)

        anchored_images: List[AnchoredImage] = []
        for row_zero, col_zero, r_embed in anchors:
            if r_embed in rel_map:
                media_path = rel_map[r_embed]
                anchored_images.append(AnchoredImage(row=row_zero, col=col_zero, media_path=media_path))

        filtered = [ai for ai in anchored_images if ai.col == img_col_idx]
        filtered.sort(key=lambda x: x.row)

        results: List[ExtractedImageDetail] = []
        for ai in filtered:
            row_number = ai.row + 1
            name_cell = ws.cell(row=row_number, column=name_col_idx + 1).value
            name_raw = str(name_cell) if name_cell is not None else f"image_row{row_number}"
            raw_bytes = zf.read(ai.media_path)
            results.append(ExtractedImageDetail(row=row_number, sheet=ws.title, name_raw=name_raw, image_bytes=raw_bytes))

            if max_images and max_images > 0 and len(results) > max_images:
                raise ValueError(f"Too many images: limit is {max_images}")

        return results


def extract_images_details_with_total(
    xlsx_bytes: bytes,
    image_col_letter: str,
    name_col_letter: str,
    soft_limit: int | None = None,
) -> Tuple[List[ExtractedImageDetail], int]:
    """Extract images similarly to extract_images_details, but return total available and
    optionally stop early at soft_limit without raising.

    Returns: (results, total_available)
    """
    img_col_idx = column_letter_to_index(image_col_letter)
    name_col_idx = column_letter_to_index(name_col_letter)

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    ws = wb.worksheets[0]

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        sheet_xml, sheet_rels = _get_first_sheet_paths(zf)
        drawing_xml_path = _find_drawing_for_sheet(zf, sheet_xml, sheet_rels)
        if not drawing_xml_path:
            return [], 0

        rel_map = _map_drawing_relations(zf, drawing_xml_path)
        anchors = _parse_anchored_images(zf, drawing_xml_path)

        anchored_images: List[AnchoredImage] = []
        for row_zero, col_zero, r_embed in anchors:
            if r_embed in rel_map:
                media_path = rel_map[r_embed]
                anchored_images.append(AnchoredImage(row=row_zero, col=col_zero, media_path=media_path))

        filtered = [ai for ai in anchored_images if ai.col == img_col_idx]
        filtered.sort(key=lambda x: x.row)

        total_available = len(filtered)

        results: List[ExtractedImageDetail] = []
        max_count = soft_limit if (soft_limit and soft_limit > 0) else total_available
        for ai in filtered[:max_count]:
            row_number = ai.row + 1
            name_cell = ws.cell(row=row_number, column=name_col_idx + 1).value
            name_raw = str(name_cell) if name_cell is not None else f"image_row{row_number}"
            raw_bytes = zf.read(ai.media_path)
            results.append(ExtractedImageDetail(row=row_number, sheet=ws.title, name_raw=name_raw, image_bytes=raw_bytes))

        return results, total_available


